import openpyxl
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException

class product:
    
    def __init__(self, title, URI, excelSheet, keys=None):
        self.title = title
        self.sheet = excelSheet
        self.URI = URI
        if keys == None or self.title not in keys:
            self.col = self.sheet.max_column+1
            self.sheet.cell(2, self.col).value = self.title
            self.sheet.cell(3, self.col).value = self.URI
        else:
            self.col = keys.index(self.title)

    # Purpose: Sets URI in case it must be updated\
    # Params:  URI - The uri of the of the item
    def setURI(self, URI):
        self.URI = URI
        self.sheet.cell(3, self.col).value = URI
    
    # Purpose: To input a price directyl
    # Usage:   Used when recreating already existing entries from excel
    def setPrice(self, price):
        self.price = price

    def getCol(self):
        return self.col

    async def shiftCol(self, originCol):
        if self.col > originCol:
            self.col-= 1

    # Gets the price in the excel sheet
    def currPrice(self):
        return self.price

    async def getPriceFromPage(self, driver=webdriver):
        driver.get(self.URI)
        delay = 3
        try: self.price = WebDriverWait(driver, delay).until(EC.presence_of_element_located((webdriver.common.by.By.CLASS_NAME, "a-price-whole"))).text
        except TimeoutException: return 1
        self.price+= "."
        try: self.price+= WebDriverWait(driver, delay).until(EC.presence_of_element_located((webdriver.common.by.By.CLASS_NAME, "a-price-fraction"))).text
        except TimeoutException: return 1
        self.sheet.cell(self.sheet.max_row, self.col).value = float(self.price.replace(",", ""))
        return

    # Returns title and URI
    def __str__(self):
        string = "Title: "
        string+= self.title
        string+= "\nURI: "
        string+= self.URI
        string+= "\n"
        return string


if __name__ == "__main__":
    excelBook = openpyxl.Workbook()
    excelBook.worksheets[0].title = "Data"
    excelBook.worksheets[0].cell(2,1).value = "Titles: "
    excelBook.worksheets[0].cell(3,1).value = "URI: "
    excelBook.worksheets[0].cell(1,2).value = "Dates"
    print(excelBook.worksheets[0].max_column)
    p1 = product("AHHHHHHH", excelBook.worksheets[0])
    p1.setURI("baltimores")
    excelBook.save(filename="aTest.xlsx")